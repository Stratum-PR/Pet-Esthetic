# Noloco Payroll Export
# Sheets: Time Entries (PAYROLL TIMESHEET), Employee Summary (BY EMPLOYEE SUMMARY), Payroll (PAY CALCULATIONS)
#
# SOURCE TABLE CHANGE (Feb 2026):
#   Previously read from: timesheets (timesheetsCollection)
#   Now reads from: Splash Page Clocks (testClockingActionCollection)
#
#   Field renames from old source → new source:
#     timesheets.clockDatetime     → clocks.clockIn
#     timesheets.clockOutDatetime  → clocks.clockOut
#     timesheets.shiftHoursWorked  → clocks.shiftHours
#     timesheets.timesheetDate     → derived from clocks.clockIn (date portion)
#     timesheets.approved          → clocks.approved  (NEW field on Clocks table)
#     timesheets.employeeFullName  → clocks.employeeFullName  (already existed)
#     timesheets.employeePin       → clocks.employeePin  (unchanged)

import os
import time
from datetime import date, datetime, timedelta
from itertools import groupby

try:
    from dotenv import load_dotenv
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    env_path = os.path.join(project_root, '.env')
    load_dotenv(env_path)
except ImportError:
    pass

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from tools import send_gmail
from config import Config


def _format_date(d):
    """Format date or YYYY-MM-DD string as MM/DD/YYYY."""
    if d is None:
        return ""
    if isinstance(d, str):
        s = (d or "")[:10]
        for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt).strftime("%m/%d/%Y")
            except ValueError:
                continue
        return d
    return d.strftime("%m/%d/%Y")


def _format_time(iso_str):
    """Format ISO datetime string as 12h time (e.g. 05:00 PM) in Puerto Rico time (AST, UTC-4)."""
    if not iso_str:
        return ""
    try:
        from datetime import timezone, timedelta
        s = str(iso_str).replace("Z", "+00:00").split(".")[0]
        # Re-add UTC offset if stripped by split
        if "+" not in s and len(str(iso_str).split(".")[0]) != len(s):
            s += "+00:00"
        dt = datetime.fromisoformat(s)
        # If datetime is naive, assume UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        # Convert to AST (UTC-4)
        ast_tz = timezone(timedelta(hours=-4))
        dt = dt.astimezone(ast_tz)
        return dt.strftime("%I:%M %p")
    except Exception:
        return ""


def _format_period(period):
    """Format period dict as 'December 07 - December 18, 2025'."""
    start = datetime.strptime(period["start_date"], "%Y-%m-%d").date()
    end = datetime.strptime(period["end_date"], "%Y-%m-%d").date()
    return f"{start.strftime('%B %d')} - {end.strftime('%B %d')}, {end.year}"


def _format_generated():
    """Format current time as 'December 17, 2025 at 05:16 PM' in Puerto Rico time."""
    from datetime import timezone, timedelta
    pr_tz = timezone(timedelta(hours=-4))
    now_pr = datetime.now(pr_tz)
    return now_pr.strftime("%B %d, %Y at %I:%M %p")


_WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def _add_logo_header(ws, logo_path):
    if logo_path and os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            img.width = 247
            img.height = 72
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 74.25
            for col in ["A", "B", "C", "D"]:
                ws[f"{col}1"].fill = _WHITE_FILL
            return True
        except Exception:
            return False
    return False


def create_time_entries_sheet(wb, company, period_formatted, generated_str, time_entry_rows, styles, logo_path=None):
    """Sheet 1: Time Entries. Unchanged from original."""
    ws = wb.create_sheet("Time Entries")
    r = 2 if _add_logo_header(ws, logo_path) else 1
    ws[f"A{r}"] = f"{company} - PAYROLL TIMESHEET"
    ws[f"A{r}"].font = styles["title_font"]
    r += 1
    ws[f"A{r}"] = f"Pay Period: {period_formatted}"
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 1
    ws[f"A{r}"] = f"Generated: {generated_str}"
    ws[f"A{r}"].font = Font(size=10)
    r += 2
    headers = ["Employee ID", "Employee Name", "Date", "Clock In", "Clock Out", "Hours", "Status", "Period Start", "Period End"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c)
        cell.value = h
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = styles["border"]
    r += 1
    start_data = r
    for row in time_entry_rows:
        ws.cell(row=r, column=1).value = row.get("employeeIdVal", "")
        ws.cell(row=r, column=2).value = row.get("employeeName", "")
        ws.cell(row=r, column=3).value = row.get("date", "")
        ws.cell(row=r, column=4).value = row.get("clockIn", "")
        ws.cell(row=r, column=5).value = row.get("clockOut", "")
        cell = ws.cell(row=r, column=6)
        cell.value = row.get("hours", 0)
        cell.number_format = "0.00"
        ws.cell(row=r, column=7).value = row.get("status", "")
        ws.cell(row=r, column=8).value = row.get("periodStart", "")
        ws.cell(row=r, column=9).value = row.get("periodEnd", "")
        r += 1
    r += 1
    ws[f"A{r}"] = "TOTAL"
    ws[f"A{r}"].font = Font(bold=True, size=11)
    cell = ws.cell(row=r, column=6)
    cell.value = f"=SUM(F{start_data}:F{r-2})" if (r - 2) >= start_data else 0
    cell.font = Font(bold=True)
    cell.number_format = "0.00"
    for col, w in [("A", 12), ("B", 22), ("C", 12), ("D", 10), ("E", 10), ("F", 8), ("G", 10), ("H", 12), ("I", 12)]:
        ws.column_dimensions[col].width = w
    return ws


def create_employee_summary_sheet(wb, company, period_formatted, time_entry_rows, styles, logo_path=None):
    """Sheet 2: Employee Summary. Unchanged from original."""
    ws = wb.create_sheet("Employee Summary")
    r = 2 if _add_logo_header(ws, logo_path) else 1
    ws[f"A{r}"] = f"{company} - BY EMPLOYEE SUMMARY"
    ws[f"A{r}"].font = styles["title_font"]
    r += 1
    ws[f"A{r}"] = f"Pay Period: {period_formatted}"
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 2
    key_fn = lambda x: (x.get("employeeIdVal"), x.get("employeeName", ""))
    sorted_rows = sorted(time_entry_rows, key=key_fn)
    for (eid, ename), rows in groupby(sorted_rows, key=key_fn):
        ws[f"A{r}"] = f"Employee: {ename} (ID: {eid})"
        ws[f"A{r}"].font = Font(bold=True, size=11)
        r += 1
        headers = ["Date", "Clock In", "Clock Out", "Hours", "Status"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c)
            cell.value = h
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
            cell.border = styles["border"]
        r += 1
        first_data = r
        for row in rows:
            ws.cell(row=r, column=1).value = row.get("date", "")
            ws.cell(row=r, column=2).value = row.get("clockIn", "")
            ws.cell(row=r, column=3).value = row.get("clockOut", "")
            cell = ws.cell(row=r, column=4)
            cell.value = row.get("hours", 0)
            cell.number_format = "0.00"
            ws.cell(row=r, column=5).value = row.get("status", "")
            r += 1
        ws[f"A{r}"] = f"Subtotal - {ename}"
        ws[f"A{r}"].font = Font(bold=True, size=10)
        cell = ws.cell(row=r, column=4)
        cell.value = f"=SUM(D{first_data}:D{r-1})" if (r - 1) >= first_data else 0
        cell.font = Font(bold=True)
        cell.number_format = "0.00"
        r += 2
    for col, w in [("A", 14), ("B", 10), ("C", 10), ("D", 8), ("E", 10)]:
        ws.column_dimensions[col].width = w
    return ws


def create_payroll_sheet(wb, df_agg, company, period_formatted, styles, logo_path=None):
    """Sheet 3: Payroll calculations. Unchanged from original."""
    ws = wb.create_sheet("Payroll")
    r = 2 if _add_logo_header(ws, logo_path) else 1
    ws[f"A{r}"] = f"{company} - PAY CALCULATIONS"
    ws[f"A{r}"].font = styles["title_font"]
    r += 1
    ws[f"A{r}"] = f"Pay Period: {period_formatted}"
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 2
    ws[f"A{r}"] = "Note: Pay rates and Commission % are editable. Gross Pay = Hours x Rate. Commission Pay = Commission % x Sales Volume."
    ws[f"A{r}"].font = Font(italic=True, size=10)
    r += 2
    headers = ["Employee ID", "Employee Name", "Total Hours", "Hourly Rate", "Gross Pay", "Commission %", "Sales Volume", "Commission Pay"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c)
        cell.value = h
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = styles["border"]
    r += 1
    start_data = r
    for _, rec in df_agg.iterrows():
        eid = rec.get("employeeIdVal", "")
        name = rec.get("employeeFullName", "Unknown")
        hours = float(rec.get("shiftHours") or 0)
        rate_val = rec.get("payRate")
        try:
            rate = float(rate_val) if rate_val is not None and str(rate_val).strip() != "" else 0.0
        except (ValueError, TypeError):
            rate = 0.0
        ws.cell(row=r, column=1).value = eid
        ws.cell(row=r, column=2).value = name
        cell = ws.cell(row=r, column=3)
        cell.value = hours
        cell.number_format = "0.00"
        cell = ws.cell(row=r, column=4)
        cell.value = rate if rate else None
        cell.number_format = "0.00"
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell = ws.cell(row=r, column=5)
        cell.value = f"=C{r}*D{r}"
        cell.number_format = "$#,##0.00"
        cell.font = Font(bold=True)
        cell = ws.cell(row=r, column=6)
        cell.value = None
        cell.number_format = "0.00%"
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell = ws.cell(row=r, column=7)
        cell.value = None
        cell.number_format = "#,##0.00"
        cell = ws.cell(row=r, column=8)
        cell.value = f"=F{r}*G{r}"
        cell.number_format = "$#,##0.00"
        cell.font = Font(bold=True)
        r += 1
    r += 1
    ws[f"A{r}"] = "TOTALS"
    ws[f"A{r}"].font = Font(bold=True, size=11)
    cell = ws.cell(row=r, column=3)
    cell.value = f"=SUM(C{start_data}:C{r-2})" if (r - 2) >= start_data else 0
    cell.font = Font(bold=True)
    cell.number_format = "0.00"
    cell = ws.cell(row=r, column=5)
    cell.value = f"=SUM(E{start_data}:E{r-2})" if (r - 2) >= start_data else 0
    cell.font = Font(bold=True)
    cell.number_format = "$#,##0.00"
    cell = ws.cell(row=r, column=8)
    cell.value = f"=SUM(H{start_data}:H{r-2})" if (r - 2) >= start_data else 0
    cell.font = Font(bold=True)
    cell.number_format = "$#,##0.00"
    for col, w in [("A", 12), ("B", 25), ("C", 12), ("D", 12), ("E", 12), ("F", 12), ("G", 14), ("H", 14)]:
        ws.column_dimensions[col].width = w
    return ws


# =============================================================================
# CONFIGURATION & API
# =============================================================================

API_TOKEN = os.getenv("NOLOCO_API_TOKEN")
PROJECT_ID = os.getenv("NOLOCO_PROJECT_ID")

if not API_TOKEN or not str(API_TOKEN).strip():
    raise Exception("ERROR: NOLOCO_API_TOKEN not set!")
if not PROJECT_ID or not str(PROJECT_ID).strip():
    raise Exception("ERROR: NOLOCO_PROJECT_ID not set!")

API_TOKEN = str(API_TOKEN).strip()
PROJECT_ID = str(PROJECT_ID).strip()

REFERENCE_MONDAY = date(2026, 1, 12)
MAX_RETRIES = 3
RETRY_DELAY = 2


def _run_graphql(api_url, headers, query, retry_count=0):
    """Execute GraphQL query with retry. Unchanged from original."""
    try:
        proxies = {"http": None, "https": None}
        resp = requests.post(api_url, headers=headers, json={"query": query}, proxies=proxies, timeout=30)
        if resp.status_code == 429 and retry_count < MAX_RETRIES:
            time.sleep(RETRY_DELAY * (retry_count + 1))
            return _run_graphql(api_url, headers, query, retry_count + 1)
        if resp.status_code >= 500 and retry_count < MAX_RETRIES:
            time.sleep(RETRY_DELAY * (retry_count + 1))
            return _run_graphql(api_url, headers, query, retry_count + 1)
        if resp.status_code == 401:
            raise Exception("Authentication failed. Check NOLOCO_API_TOKEN.")
        if resp.status_code != 200:
            raise Exception(f"API error: {resp.status_code} - {resp.text[:300]}")
        data = resp.json()
        if "errors" in data:
            msgs = [e.get("message", "?") for e in data["errors"]]
            raise Exception("GraphQL error: " + "; ".join(msgs))
        return data.get("data") or {}
    except requests.exceptions.Timeout:
        if retry_count < MAX_RETRIES:
            time.sleep(RETRY_DELAY * (retry_count + 1))
            return _run_graphql(api_url, headers, query, retry_count + 1)
        raise
    except requests.exceptions.ConnectionError:
        if retry_count < MAX_RETRIES:
            time.sleep(RETRY_DELAY * (retry_count + 1))
            return _run_graphql(api_url, headers, query, retry_count + 1)
        raise


def _pay_period_for(target_date):
    """Bi-weekly pay period (Mon–Sun, 14 days). Unchanged from original."""
    monday = target_date - timedelta(days=target_date.weekday())
    days_from_ref = (monday - REFERENCE_MONDAY).days
    period_num = days_from_ref // 14
    start = REFERENCE_MONDAY + timedelta(days=period_num * 14)
    end = start + timedelta(days=13)
    return {"start_date": start.strftime("%Y-%m-%d"), "end_date": end.strftime("%Y-%m-%d")}


def _is_approved(rec):
    v = rec.get("approved")
    if v is True:
        return True
    if isinstance(v, str) and (v or "").strip().lower() == "true":
        return True
    return False


def _fetch_clock_records(api_url, headers):
    """
    Fetch all Splash Page Clock records for the export.

    CHANGED FROM ORIGINAL:
      Was: _fetch_timesheets() → timesheetsCollection
           fields: employeePin, employeeFullName, timesheetDate,
                   approved, shiftHoursWorked, clockDatetime, clockOutDatetime
      Now: _fetch_clock_records() → testClockingActionCollection
           fields: employeePin, employeeFullName, clockIn (→ date),
                   approved, shiftHours, clockIn, clockOut
    """
    out = []
    cursor = None
    while True:
        if cursor:
            q = f'query {{ testClockingActionCollection(first: 100, after: "{cursor}") {{ edges {{ node {{ id employeePin employeeFullName clockIn clockOut approved shiftHours }} }} pageInfo {{ hasNextPage endCursor }} }} }}'
        else:
            q = "query { testClockingActionCollection(first: 100) { edges { node { id employeePin employeeFullName clockIn clockOut approved shiftHours } } pageInfo { hasNextPage endCursor } } }"
        data = _run_graphql(api_url, headers, q)
        coll = data.get("testClockingActionCollection") or {}
        edges = coll.get("edges") or []
        pi = coll.get("pageInfo") or {}
        for e in edges:
            n = e.get("node") or {}
            clock_in = n.get("clockIn") or ""
            # Derive date from clockIn — same role as timesheetDate
            if "T" in clock_in:
                timesheet_date = clock_in.split("T")[0]
            elif "," in clock_in:
                # Handle Noloco format: "2/14/2026, 8:48 AM"
                timesheet_date = clock_in.split(",")[0].strip()
            else:
                timesheet_date = clock_in
            out.append({
                "id": n.get("id"),
                "employeePin": n.get("employeePin"),
                "employeeFullName": n.get("employeeFullName"),
                "timesheetDate": timesheet_date,       # derived from clockIn
                "approved": n.get("approved"),
                "shiftHoursWorked": n.get("shiftHours"),   # renamed
                "clockDatetime": n.get("clockIn"),          # renamed
                "clockOutDatetime": n.get("clockOut"),      # renamed
            })
        if not pi.get("hasNextPage"):
            break
        cursor = pi.get("endCursor")
        if not cursor:
            break
    return out


def _fetch_employees(api_url, headers):
    """Fetch employee pay rates. Unchanged from original."""
    out = {}
    cursor = None
    while True:
        if cursor:
            q = f'query {{ employeesCollection(first: 100, after: "{cursor}") {{ edges {{ node {{ employeeIdVal payRate }} }} pageInfo {{ hasNextPage endCursor }} }} }}'
        else:
            q = "query { employeesCollection(first: 100) { edges { node { employeeIdVal payRate } } pageInfo { hasNextPage endCursor } } }"
        data = _run_graphql(api_url, headers, q)
        coll = data.get("employeesCollection") or {}
        edges = coll.get("edges") or []
        pi = coll.get("pageInfo") or {}
        for e in edges:
            n = e.get("node") or {}
            eid = n.get("employeeIdVal")
            if eid is not None:
                key = str(eid).strip()
                out[key] = {"payRate": n.get("payRate")}
        if not pi.get("hasNextPage"):
            break
        cursor = pi.get("endCursor")
        if not cursor:
            break
    return out


def _upload_to_noloco(api_url: str, headers: dict, file_path: str, period: dict) -> bool:
    """
    Upload the Excel file to Noloco's file storage and create a record in the
    documents table.  Noloco accepts multipart uploads via the REST endpoint:
        POST https://api.portals.noloco.io/files/{PROJECT_ID}
    The returned file URL is then written to the documents table via GraphQL.
    """
    upload_url = f"https://api.portals.noloco.io/files/{PROJECT_ID}"
    filename = os.path.basename(file_path)

    # Step 1 — upload the binary file
    try:
        with open(file_path, "rb") as f:
            upload_headers = {"Authorization": f"Bearer {API_TOKEN}"}
            resp = requests.post(
                upload_url,
                headers=upload_headers,
                files={"file": (filename, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=60,
            )
        if resp.status_code not in (200, 201):
            print(f"  ⚠️  File upload failed: HTTP {resp.status_code} — {resp.text[:200]}")
            return False

        upload_data = resp.json()
        file_url = upload_data.get("url") or upload_data.get("Location") or upload_data.get("location")
        if not file_url:
            print(f"  ⚠️  Upload succeeded but no URL in response: {upload_data}")
            return False

        print(f"  ✓ File uploaded: {file_url}")

    except Exception as e:
        print(f"  ⚠️  Upload request failed: {e}")
        return False

    # Step 2 — create a record in the documents table linking to the file
    period_label = f"{period['start_date']} to {period['end_date']}"
    mutation = f"""
    mutation {{
      createDocument(
        name: "{filename}",
        periodLabel: "{period_label}",
        file: {{ url: "{file_url}", filename: "{filename}" }}
      ) {{
        id
      }}
    }}
    """
    try:
        result = _run_graphql(api_url, headers, mutation)
        doc_id = (result.get("createDocument") or {}).get("id")
        if doc_id:
            print(f"  ✓ Noloco document record created: ID {doc_id}")
            return True
        else:
            print(f"  ⚠️  Document mutation returned no ID: {result}")
            return False
    except Exception as e:
        print(f"  ⚠️  Document record creation failed: {e}")
        return False


def run_export():
    api_url = f"https://api.portals.noloco.io/data/{PROJECT_ID}"
    headers = {"Authorization": f"Bearer {API_TOKEN}", "Content-Type": "application/json"}

    today = date.today()
    today = date(2026, 5, 4)  # Last Monday — override instead of date.today()
    period = _pay_period_for(today)
    period_start_date = datetime.strptime(period["start_date"], "%Y-%m-%d").date()

    # If today is the start of a new period, use the previous period
    if today == period_start_date:
        prev_period_start = period_start_date - timedelta(days=14)
        prev_period_end = prev_period_start + timedelta(days=13)
        period = {
            "start_date": prev_period_start.strftime("%Y-%m-%d"),
            "end_date": prev_period_end.strftime("%Y-%m-%d")
        }

    period_start = datetime.strptime(period["start_date"], "%Y-%m-%d").date()
    period_end = datetime.strptime(period["end_date"], "%Y-%m-%d").date()

    print("Noloco Payroll Export")
    print("Source: Splash Page Clocks (testClockingActionCollection)")
    print("=" * 60)
    print(f"Pay period: {period['start_date']} to {period['end_date']}")
    print("Fetching clock records...")

    # CHANGED: was _fetch_timesheets, now _fetch_clock_records
    all_records = _fetch_clock_records(api_url, headers)
    print("Fetching employees...")
    emp_map = _fetch_employees(api_url, headers)

    # Filter: in period
    # Internal field names (timesheetDate, shiftHoursWorked, clockDatetime, clockOutDatetime)
    # are already normalized in _fetch_clock_records, so no changes needed below this line.
    time_entry_rows = []
    rows = []
    for rec in all_records:
        td = (rec.get("timesheetDate") or "").split("T")[0]
        if not td:
            continue
        # Handle both YYYY-MM-DD and M/D/YYYY
        d = None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
            try:
                d = datetime.strptime(td, fmt).date()
                break
            except ValueError:
                continue
        if not d or not (period_start <= d <= period_end):
            continue

        pin = rec.get("employeePin")
        if pin is None:
            continue
        key = str(pin).strip()
        emp = emp_map.get(key) or {}
        employee_name = rec.get("employeeFullName") or "Unknown"

        time_entry_rows.append({
            "employeeIdVal": pin,
            "employeeName": employee_name,
            "date": _format_date(td),
            "clockIn": _format_time(rec.get("clockDatetime")),
            "clockOut": _format_time(rec.get("clockOutDatetime")),
            "hours": rec.get("shiftHoursWorked") or 0,
            "status": "Approved" if _is_approved(rec) else "Pending",
            "periodStart": _format_date(period["start_date"]),
            "periodEnd": _format_date(period["end_date"]),
        })
        rows.append({
            "employeeIdVal": pin,
            "employeeFullName": employee_name,
            "shiftHours": rec.get("shiftHoursWorked") or 0,
            "payRate": emp.get("payRate"),
        })

    df_agg = pd.DataFrame(rows).groupby("employeeIdVal", as_index=False).agg(
        employeeFullName=("employeeFullName", "first"),
        shiftHours=("shiftHours", "sum"),
        payRate=("payRate", "first"),
    ) if rows else pd.DataFrame(columns=["employeeIdVal", "employeeFullName", "shiftHours", "payRate"])

    if len(time_entry_rows) == 0:
        print("No approved clock records in this pay period; export will have empty sheets.")

    company = os.getenv("COMPANY_NAME", "Pet Esthetic")
    period_formatted = _format_period(period)
    generated_str = _format_generated()

    thin = Side(style="thin")
    styles = {
        "title_font": Font(bold=True, size=14),
        "header_font": Font(bold=True),
        "header_fill": PatternFill(start_color="F88379", end_color="F88379", fill_type="solid"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }

    _script_dir = os.path.dirname(os.path.abspath(__file__))
    logo_path = os.path.abspath(os.path.join(_script_dir, "..", "assets", "pet_esthetic_transparent.png"))
    if not os.path.exists(logo_path):
        logo_path = None

    wb = Workbook()
    create_time_entries_sheet(wb, company, period_formatted, generated_str, time_entry_rows, styles, logo_path)
    create_employee_summary_sheet(wb, company, period_formatted, time_entry_rows, styles, logo_path)
    create_payroll_sheet(wb, df_agg, company, period_formatted, styles, logo_path)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    out_path = f"Payroll_Export_{period['start_date']}_to_{period['end_date']}.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")

    # Upload to Noloco documents table
    try:
        print("\nUploading to Noloco documents...")
        noloco_url = f"https://api.portals.noloco.io/data/{PROJECT_ID}"
        noloco_headers = {"Authorization": f"Bearer {API_TOKEN}", "Content-Type": "application/json"}
        uploaded = _upload_to_noloco(noloco_url, noloco_headers, out_path, period)
        if not uploaded:
            print("  Upload skipped or failed — file still saved locally and will be emailed.")
    except Exception as e:
        print(f"  Warning: Noloco upload error: {e}")

    # Send email — unchanged from original
    try:
        config = Config.from_env()
        if config.email_recipients:
            print("\nSending payroll export via email...")
            subject = f"Payroll Export - {period_formatted}"
            body_html = f"""
            <html>
            <body>
                <h2>Payroll Export Report</h2>
                <p>Please find attached the payroll export for the period:</p>
                <p><strong>{period_formatted}</strong></p>
                <p>Generated: {generated_str}</p>
                <p>This file contains three sheets:</p>
                <ul>
                    <li><strong>Time Entries</strong> - Detailed clock entries</li>
                    <li><strong>Employee Summary</strong> - Summary by employee</li>
                    <li><strong>Payroll</strong> - Pay calculations with editable rates</li>
                </ul>
            </body>
            </html>
            """
            send_gmail(
                to_emails=config.email_recipients,
                subject=subject,
                body_html=body_html,
                attachment_path=out_path,
                attachment_filename=os.path.basename(out_path)
            )
            print(f"✓ Email sent successfully to: {', '.join(config.email_recipients)}")
        else:
            print("\n⚠️  Email recipients not configured. Skipping email.")
    except Exception as e:
        print(f"\n⚠️  Warning: Failed to send email: {str(e)}")
        print("  Export file saved successfully.")

    print("Done.")


if __name__ == "__main__":
    try:
        run_export()
    except KeyboardInterrupt:
        print("\nInterrupted.")
        exit(130)
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
