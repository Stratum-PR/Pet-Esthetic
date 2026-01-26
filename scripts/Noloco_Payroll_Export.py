# Noloco Payroll Export: matches "Pet Esthetic Payroll Report" format
# Sheets: Time Entries (PAYROLL TIMESHEET), Employee Summary (BY EMPLOYEE SUMMARY), Payroll (PAY CALCULATIONS)

import os
import time
from datetime import date, datetime, timedelta
from itertools import groupby

try:
    from dotenv import load_dotenv
    # Load .env from project root (parent directory of scripts folder)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    env_path = os.path.join(project_root, '.env')
    load_dotenv(env_path)
except ImportError:
    pass

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

# Import email functions from timesheets script
from tools import send_gmail
from config import Config


def _format_date(d):
    """Format date or YYYY-MM-DD string as MM/DD/YYYY."""
    if d is None:
        return ""
    if isinstance(d, str):
        d = datetime.strptime((d or "")[:10], "%Y-%m-%d").date()
    return d.strftime("%m/%d/%Y")


def _format_time(iso_str):
    """Format ISO datetime string as 12h time (e.g. 05:00 PM)."""
    if not iso_str:
        return ""
    try:
        s = str(iso_str).replace("Z", "+00:00").split(".")[0]
        dt = datetime.fromisoformat(s)
        return dt.strftime("%I:%M %p")
    except Exception:
        return ""


def _format_period(period):
    """Format period dict as 'December 07 - December 18, 2025'."""
    start = datetime.strptime(period["start_date"], "%Y-%m-%d").date()
    end = datetime.strptime(period["end_date"], "%Y-%m-%d").date()
    return f"{start.strftime('%B %d')} - {end.strftime('%B %d')}, {end.year}"


def _format_generated():
    """Format current time as 'December 17, 2025 at 05:16 PM'."""
    return datetime.now().strftime("%B %d, %Y at %I:%M %p")


_WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def _add_logo_header(ws, logo_path):
    """Add Pet Esthetic logo at A1 as sheet header. Returns True if added (content starts row 2), else False (row 1).
    Size and row 1 height match the Time Entries tab setup (247×72 px, row 1 = 74.25 pt). Applied to all tabs.
    Cells A1:D1 are filled white behind the logo."""
    if logo_path and os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            # Matches Time Entries: 247×72 px (~3.4:1), row 1 = 74.25 pt
            img.width = 247
            img.height = 72
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 74.25
            for col in ["A", "B", "C", "D"]:
                ws[f"{col}1"].fill = _WHITE_FILL
            return True
        except Exception:
            return False
    return False


def create_time_entries_sheet(wb, company, period_formatted, generated_str, time_entry_rows, styles, logo_path=None):
    """Sheet 1: Time Entries (PAYROLL TIMESHEET). One row per timesheet."""
    ws = wb.create_sheet("Time Entries")
    r = 2 if _add_logo_header(ws, logo_path) else 1
    ws[f"A{r}"] = f"{company} - PAYROLL TIMESHEET"
    ws[f"A{r}"].font = styles["title_font"]
    r += 1
    ws[f"A{r}"] = f"Pay Period: {period_formatted}"
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 1
    ws[f"A{r}"] = f"Generated: {generated_str}"
    ws[f"A{r}"].font = Font(size=10)
    r += 2
    headers = ["Employee ID", "Employee Name", "Date", "Clock In", "Clock Out", "Hours", "Status", "Period Start", "Period End"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c)
        cell.value = h
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = styles["border"]
    r += 1
    start_data = r
    for row in time_entry_rows:
        ws.cell(row=r, column=1).value = row.get("employeeIdVal", "")
        ws.cell(row=r, column=2).value = row.get("employeeName", "")
        ws.cell(row=r, column=3).value = row.get("date", "")
        ws.cell(row=r, column=4).value = row.get("clockIn", "")
        ws.cell(row=r, column=5).value = row.get("clockOut", "")
        cell = ws.cell(row=r, column=6)
        cell.value = row.get("hours", 0)
        cell.number_format = "0.00"
        ws.cell(row=r, column=7).value = row.get("status", "")
        ws.cell(row=r, column=8).value = row.get("periodStart", "")
        ws.cell(row=r, column=9).value = row.get("periodEnd", "")
        r += 1
    r += 1
    ws[f"A{r}"] = "TOTAL"
    ws[f"A{r}"].font = Font(bold=True, size=11)
    cell = ws.cell(row=r, column=6)
    cell.value = f"=SUM(F{start_data}:F{r-2})" if (r - 2) >= start_data else 0
    cell.font = Font(bold=True)
    cell.number_format = "0.00"
    for col, w in [("A", 12), ("B", 22), ("C", 12), ("D", 10), ("E", 10), ("F", 8), ("G", 10), ("H", 12), ("I", 12)]:
        ws.column_dimensions[col].width = w
    return ws


def create_employee_summary_sheet(wb, company, period_formatted, time_entry_rows, styles, logo_path=None):
    """Sheet 2: Employee Summary (BY EMPLOYEE SUMMARY). One block per employee."""
    ws = wb.create_sheet("Employee Summary")
    r = 2 if _add_logo_header(ws, logo_path) else 1
    ws[f"A{r}"] = f"{company} - BY EMPLOYEE SUMMARY"
    ws[f"A{r}"].font = styles["title_font"]
    r += 1
    ws[f"A{r}"] = f"Pay Period: {period_formatted}"
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 2
    key_fn = lambda x: (x.get("employeeIdVal"), x.get("employeeName", ""))
    sorted_rows = sorted(time_entry_rows, key=key_fn)
    for (eid, ename), rows in groupby(sorted_rows, key=key_fn):
        ws[f"A{r}"] = f"Employee: {ename} (ID: {eid})"
        ws[f"A{r}"].font = Font(bold=True, size=11)
        r += 1
        headers = ["Date", "Clock In", "Clock Out", "Hours", "Status"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c)
            cell.value = h
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
            cell.border = styles["border"]
        r += 1
        first_data = r
        for row in rows:
            ws.cell(row=r, column=1).value = row.get("date", "")
            ws.cell(row=r, column=2).value = row.get("clockIn", "")
            ws.cell(row=r, column=3).value = row.get("clockOut", "")
            cell = ws.cell(row=r, column=4)
            cell.value = row.get("hours", 0)
            cell.number_format = "0.00"
            ws.cell(row=r, column=5).value = row.get("status", "")
            r += 1
        ws[f"A{r}"] = f"Subtotal - {ename}"
        ws[f"A{r}"].font = Font(bold=True, size=10)
        cell = ws.cell(row=r, column=4)
        cell.value = f"=SUM(D{first_data}:D{r-1})" if (r - 1) >= first_data else 0
        cell.font = Font(bold=True)
        cell.number_format = "0.00"
        r += 2
    for col, w in [("A", 14), ("B", 10), ("C", 10), ("D", 8), ("E", 10)]:
        ws.column_dimensions[col].width = w
    return ws


def create_payroll_sheet(wb, df_agg, company, period_formatted, styles, logo_path=None):
    """Sheet 3: Payroll (PAY CALCULATIONS). Employee ID, Name, Total Hours, Hourly Rate (editable), Gross Pay, Commission % (editable), Sales Volume, Commission Pay."""
    ws = wb.create_sheet("Payroll")
    r = 2 if _add_logo_header(ws, logo_path) else 1
    ws[f"A{r}"] = f"{company} - PAY CALCULATIONS"
    ws[f"A{r}"].font = styles["title_font"]
    r += 1
    ws[f"A{r}"] = f"Pay Period: {period_formatted}"
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 2
    ws[f"A{r}"] = "Note: Pay rates and Commission % are editable. Gross Pay = Hours x Rate. Commission Pay = Commission % x Sales Volume."
    ws[f"A{r}"].font = Font(italic=True, size=10)
    r += 2
    headers = ["Employee ID", "Employee Name", "Total Hours", "Hourly Rate", "Gross Pay", "Commission %", "Sales Volume", "Commission Pay"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c)
        cell.value = h
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = styles["border"]
    r += 1
    start_data = r
    for _, rec in df_agg.iterrows():
        eid = rec.get("employeeIdVal", "")
        name = rec.get("users_fullName", "Unknown")
        hours = float(rec.get("shiftHoursWorked") or 0)
        rate_val = rec.get("users_payRate")
        try:
            rate = float(rate_val) if rate_val is not None and str(rate_val).strip() != "" else 0.0
        except (ValueError, TypeError):
            rate = 0.0
        ws.cell(row=r, column=1).value = eid
        ws.cell(row=r, column=2).value = name
        cell = ws.cell(row=r, column=3)
        cell.value = hours
        cell.number_format = "0.00"
        cell = ws.cell(row=r, column=4)
        cell.value = rate if rate else None
        cell.number_format = "0.00"
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell = ws.cell(row=r, column=5)
        cell.value = f"=C{r}*D{r}"
        cell.number_format = "$#,##0.00"
        cell.font = Font(bold=True)
        # Commission % (editable, gray), Sales Volume (user entry), Commission Pay = F*G
        cell = ws.cell(row=r, column=6)
        cell.value = None
        cell.number_format = "0.00%"
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell = ws.cell(row=r, column=7)
        cell.value = None
        cell.number_format = "#,##0.00"
        cell = ws.cell(row=r, column=8)
        cell.value = f"=F{r}*G{r}"
        cell.number_format = "$#,##0.00"
        cell.font = Font(bold=True)
        r += 1
    r += 1
    ws[f"A{r}"] = "TOTALS"
    ws[f"A{r}"].font = Font(bold=True, size=11)
    cell = ws.cell(row=r, column=3)
    cell.value = f"=SUM(C{start_data}:C{r-2})" if (r - 2) >= start_data else 0
    cell.font = Font(bold=True)
    cell.number_format = "0.00"
    cell = ws.cell(row=r, column=5)
    cell.value = f"=SUM(E{start_data}:E{r-2})" if (r - 2) >= start_data else 0
    cell.font = Font(bold=True)
    cell.number_format = "$#,##0.00"
    cell = ws.cell(row=r, column=8)
    cell.value = f"=SUM(H{start_data}:H{r-2})" if (r - 2) >= start_data else 0
    cell.font = Font(bold=True)
    cell.number_format = "$#,##0.00"
    for col, w in [("A", 12), ("B", 25), ("C", 12), ("D", 12), ("E", 12), ("F", 12), ("G", 14), ("H", 14)]:
        ws.column_dimensions[col].width = w
    return ws




# =============================================================================
# CONFIGURATION & API (for run when executed as script)
# =============================================================================

API_TOKEN = os.getenv("NOLOCO_API_TOKEN")
PROJECT_ID = os.getenv("NOLOCO_PROJECT_ID")

if not API_TOKEN or not str(API_TOKEN).strip():
    raise Exception("ERROR: NOLOCO_API_TOKEN not set!")
if not PROJECT_ID or not str(PROJECT_ID).strip():
    raise Exception("ERROR: NOLOCO_PROJECT_ID not set!")

API_TOKEN = str(API_TOKEN).strip()
PROJECT_ID = str(PROJECT_ID).strip()

REFERENCE_MONDAY = date(2026, 1, 12)  # Matches Noloco_Add_Payroll_Records
MAX_RETRIES = 3
RETRY_DELAY = 2


def _run_graphql(api_url, headers, query, retry_count=0):
    """Execute GraphQL query with retry. Uses api_url and headers from env."""
    try:
        proxies = {"http": None, "https": None}
        resp = requests.post(
            api_url,
            headers=headers,
            json={"query": query},
            proxies=proxies,
            timeout=30,
        )
        if resp.status_code == 429 and retry_count < MAX_RETRIES:
            time.sleep(RETRY_DELAY * (retry_count + 1))
            return _run_graphql(api_url, headers, query, retry_count + 1)
        if resp.status_code >= 500 and retry_count < MAX_RETRIES:
            time.sleep(RETRY_DELAY * (retry_count + 1))
            return _run_graphql(api_url, headers, query, retry_count + 1)
        if resp.status_code == 401:
            raise Exception("Authentication failed. Check NOLOCO_API_TOKEN.")
        if resp.status_code != 200:
            raise Exception(f"API error: {resp.status_code} - {resp.text[:300]}")
        data = resp.json()
        if "errors" in data:
            msgs = [e.get("message", "?") for e in data["errors"]]
            raise Exception("GraphQL error: " + "; ".join(msgs))
        return data.get("data") or {}
    except requests.exceptions.Timeout:
        if retry_count < MAX_RETRIES:
            time.sleep(RETRY_DELAY * (retry_count + 1))
            return _run_graphql(api_url, headers, query, retry_count + 1)
        raise
    except requests.exceptions.ConnectionError:
        if retry_count < MAX_RETRIES:
            time.sleep(RETRY_DELAY * (retry_count + 1))
            return _run_graphql(api_url, headers, query, retry_count + 1)
        raise


def _pay_period_for(target_date):
    """Bi-weekly pay period (Mon–Sun, 14 days). Matches Noloco_Add_Payroll_Records."""
    monday = target_date - timedelta(days=target_date.weekday())
    days_from_ref = (monday - REFERENCE_MONDAY).days
    period_num = days_from_ref // 14
    start = REFERENCE_MONDAY + timedelta(days=period_num * 14)
    end = start + timedelta(days=13)
    return {"start_date": start.strftime("%Y-%m-%d"), "end_date": end.strftime("%Y-%m-%d")}


def _is_approved(ts):
    v = ts.get("approved")
    if v is True:
        return True
    if isinstance(v, str) and (v or "").strip().lower() == "true":
        return True
    return False


def _fetch_timesheets(api_url, headers):
    out = []
    cursor = None
    while True:
        if cursor:
            q = f'query {{ timesheetsCollection(first: 100, after: "{cursor}") {{ edges {{ node {{ id employeePin employeeFullName timesheetDate approved shiftHoursWorked clockDatetime clockOutDatetime }} }} pageInfo {{ hasNextPage endCursor }} }} }}'
        else:
            q = "query { timesheetsCollection(first: 100) { edges { node { id employeePin employeeFullName timesheetDate approved shiftHoursWorked clockDatetime clockOutDatetime } } pageInfo { hasNextPage endCursor } } }"
        data = _run_graphql(api_url, headers, q)
        coll = data.get("timesheetsCollection") or {}
        edges = coll.get("edges") or []
        pi = coll.get("pageInfo") or {}
        for e in edges:
            n = e.get("node") or {}
            out.append({
                "id": n.get("id"),
                "employeePin": n.get("employeePin"),
                "employeeFullName": n.get("employeeFullName"),
                "timesheetDate": n.get("timesheetDate"),
                "approved": n.get("approved"),
                "shiftHoursWorked": n.get("shiftHoursWorked") or 0,
                "clockDatetime": n.get("clockDatetime"),
                "clockOutDatetime": n.get("clockOutDatetime"),
            })
        if not pi.get("hasNextPage"):
            break
        cursor = pi.get("endCursor")
        if not cursor:
            break
    return out


def _upload_file_to_noloco(api_url, headers, file_path):
    """Upload a file to Noloco's media storage and return the file URL."""
    if not os.path.exists(file_path):
        raise Exception(f"File not found: {file_path}")
    
    # Noloco file upload endpoint
    upload_url = f"https://api.portals.noloco.io/media/{PROJECT_ID}/upload"
    
    # Read file
    with open(file_path, 'rb') as f:
        file_data = f.read()
    
    # Prepare multipart form data
    filename = os.path.basename(file_path)
    files = {
        'file': (filename, file_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    }
    
    # Upload file
    upload_headers = {
        "Authorization": f"Bearer {API_TOKEN}"
    }
    
    proxies = {"http": None, "https": None}
    response = requests.post(
        upload_url,
        headers=upload_headers,
        files=files,
        proxies=proxies,
        timeout=60
    )
    
    if response.status_code != 200:
        raise Exception(f"File upload failed: {response.status_code} - {response.text[:300]}")
    
    result = response.json()
    file_url = result.get("url") or result.get("fileUrl") or result.get("file_url")
    
    if not file_url:
        # Try to extract URL from response
        response_text = response.text
        if "url" in response_text.lower():
            import json
            try:
                data = json.loads(response_text)
                file_url = data.get("url") or data.get("fileUrl") or data.get("file_url")
            except:
                pass
        
        if not file_url:
            raise Exception(f"File upload succeeded but no URL returned. Response: {response.text[:300]}")
    
    return file_url


def upload_to_noloco_documents(api_url, headers, file_path, period_formatted, period):
    """Upload Excel file to Noloco documents table."""
    if not os.path.exists(file_path):
        raise Exception(f"File not found: {file_path}")
    
    # Upload file first to get URL
    print("  Uploading file to Noloco media storage...")
    file_url = _upload_file_to_noloco(api_url, headers, file_path)
    print(f"  ✓ File uploaded, URL: {file_url[:80]}...")
    
    # Prepare document data
    filename = os.path.basename(file_path)
    document_type = "Payroll Export"
    notes = f"Payroll export for period {period_formatted} (Generated: {_format_generated()})"
    
    # Escape special characters for GraphQL strings
    def escape_graphql_string(s):
        if s is None:
            return ""
        s = str(s).replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n').replace('\r', '\\r')
        return s
    
    escaped_notes = escape_graphql_string(notes)
    escaped_filename = escape_graphql_string(filename)
    escaped_file_url = escape_graphql_string(file_url)
    
    # Create document record
    mutation = f"""
    mutation {{
        createDocuments(
            documentType: "{document_type}"
            notes: "{escaped_notes}"
            documentName: "{escaped_filename}"
            document: "{escaped_file_url}"
        ) {{
            id
        }}
    }}
    """
    
    print("  Creating document record in Noloco...")
    data = _run_graphql(api_url, headers, mutation)
    
    doc_id = data.get("createDocuments", {}).get("id")
    if not doc_id:
        raise Exception("Document creation failed - no ID returned")
    
    print(f"  ✓ Document record created with ID: {doc_id}")
    return doc_id


def _fetch_employees(api_url, headers):
    """Returns dict keyed by normalized employeeIdVal: { payRate }.
    Only fetches pay rates since employeeFullName comes from timesheets."""
    out = {}
    cursor = None
    while True:
        if cursor:
            q = f'query {{ employeesCollection(first: 100, after: "{cursor}") {{ edges {{ node {{ employeeIdVal payRate }} }} pageInfo {{ hasNextPage endCursor }} }} }}'
        else:
            q = "query { employeesCollection(first: 100) { edges { node { employeeIdVal payRate } } pageInfo { hasNextPage endCursor } } }"
        data = _run_graphql(api_url, headers, q)
        coll = data.get("employeesCollection") or {}
        edges = coll.get("edges") or []
        pi = coll.get("pageInfo") or {}
        for e in edges:
            n = e.get("node") or {}
            eid = n.get("employeeIdVal")
            if eid is not None:
                key = str(eid).strip()
                out[key] = {
                    "payRate": n.get("payRate"),
                }
        if not pi.get("hasNextPage"):
            break
        cursor = pi.get("endCursor")
        if not cursor:
            break
    return out


def run_export():
    api_url = f"https://api.portals.noloco.io/data/{PROJECT_ID}"
    headers = {"Authorization": f"Bearer {API_TOKEN}", "Content-Type": "application/json"}

    # Calculate period for today
    today = date.today()
    period = _pay_period_for(today)
    period_start_date = datetime.strptime(period["start_date"], "%Y-%m-%d").date()
    
    # Special case: If today is the start date of a new period (i.e., the day after
    # the previous period ended), use the previous period instead. This allows
    # managers to validate timesheets on the day after a period ends.
    if today == period_start_date:
        # Calculate previous period by going back 14 days from the start
        prev_period_start = period_start_date - timedelta(days=14)
        prev_period_end = prev_period_start + timedelta(days=13)
        period = {
            "start_date": prev_period_start.strftime("%Y-%m-%d"),
            "end_date": prev_period_end.strftime("%Y-%m-%d")
        }
    
    period_start = datetime.strptime(period["start_date"], "%Y-%m-%d").date()
    period_end = datetime.strptime(period["end_date"], "%Y-%m-%d").date()

    print("Noloco Payroll Export")
    print("=" * 60)
    print(f"Pay period: {period['start_date']} to {period['end_date']}")
    print("Fetching timesheets...")
    all_ts = _fetch_timesheets(api_url, headers)
    print("Fetching employees...")
    emp_map = _fetch_employees(api_url, headers)

    # Filter: in period and approved; build time_entry_rows and rows for aggregation
    time_entry_rows = []
    rows = []
    for ts in all_ts:
        if not _is_approved(ts):
            continue
        td = (ts.get("timesheetDate") or "").split("T")[0]
        if not td:
            continue
        try:
            d = datetime.strptime(td, "%Y-%m-%d").date()
        except ValueError:
            continue
        if not (period_start <= d <= period_end):
            continue
        pin = ts.get("employeePin")
        if pin is None:
            continue
        key = str(pin).strip()
        emp = emp_map.get(key) or {}
        # Use employeeFullName directly from timesheet (no matching needed)
        employee_name = ts.get("employeeFullName") or "Unknown"
        time_entry_rows.append({
            "employeeIdVal": pin,
            "employeeName": employee_name,
            "date": _format_date(td),
            "clockIn": _format_time(ts.get("clockDatetime")),
            "clockOut": _format_time(ts.get("clockOutDatetime")),
            "hours": ts.get("shiftHoursWorked") or 0,
            "status": "Approved" if _is_approved(ts) else "Pending",
            "periodStart": _format_date(period["start_date"]),
            "periodEnd": _format_date(period["end_date"]),
        })
        rows.append({
            "employeeIdVal": pin,
            "users_fullName": employee_name,
            "shiftHoursWorked": ts.get("shiftHoursWorked") or 0,
            "users_payRate": emp.get("payRate"),
        })

    df_agg = pd.DataFrame(rows).groupby("employeeIdVal", as_index=False).agg(
        users_fullName=("users_fullName", "first"),
        shiftHoursWorked=("shiftHoursWorked", "sum"),
        users_payRate=("users_payRate", "first"),
    ) if rows else pd.DataFrame(columns=["employeeIdVal", "users_fullName", "shiftHoursWorked", "users_payRate"])
    if len(time_entry_rows) == 0:
        print("No approved timesheets in this pay period; export will have empty sheets.")

    company = os.getenv("COMPANY_NAME", "Pet Esthetic")
    period_formatted = _format_period(period)
    generated_str = _format_generated()

    thin = Side(style="thin")
    styles = {
        "title_font": Font(bold=True, size=14),
        "header_font": Font(bold=True),
        "header_fill": PatternFill(start_color="F88379", end_color="F88379", fill_type="solid"),  # #f88379
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }

    _script_dir = os.path.dirname(os.path.abspath(__file__))
    logo_path = os.path.abspath(os.path.join(_script_dir, "..", "assets", "pet_esthetic_transparent.png"))
    if not os.path.exists(logo_path):
        logo_path = None

    wb = Workbook()
    create_time_entries_sheet(wb, company, period_formatted, generated_str, time_entry_rows, styles, logo_path)
    create_employee_summary_sheet(wb, company, period_formatted, time_entry_rows, styles, logo_path)
    create_payroll_sheet(wb, df_agg, company, period_formatted, styles, logo_path)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    out_path = f"Payroll_Export_{period['start_date']}_to_{period['end_date']}.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    
    # Send email with payroll export file
    try:
        config = Config.from_env()
        if config.email_recipients:
            print("\nSending payroll export via email...")
            subject = f"Payroll Export - {period_formatted}"
            body_html = f"""
            <html>
            <body>
                <h2>Payroll Export Report</h2>
                <p>Please find attached the payroll export for the period:</p>
                <p><strong>{period_formatted}</strong></p>
                <p>Generated: {generated_str}</p>
                <p>This file contains three sheets:</p>
                <ul>
                    <li><strong>Time Entries</strong> - Detailed timesheet entries</li>
                    <li><strong>Employee Summary</strong> - Summary by employee</li>
                    <li><strong>Payroll</strong> - Pay calculations with editable rates</li>
                </ul>
            </body>
            </html>
            """
            
            send_gmail(
                to_emails=config.email_recipients,
                subject=subject,
                body_html=body_html,
                attachment_path=out_path,
                attachment_filename=os.path.basename(out_path)
            )
            print(f"✓ Email sent successfully to: {', '.join(config.email_recipients)}")
        else:
            print("\n⚠️  Email recipients not configured (EMAIL_RECIPIENTS not set). Skipping email.")
    except Exception as e:
        print(f"\n⚠️  Warning: Failed to send email: {str(e)}")
        print("  The export file was saved successfully, but the email notification failed.")
    
    # Upload file to Noloco documents table
    try:
        print("\nUploading payroll export to Noloco documents table...")
        upload_to_noloco_documents(api_url, headers, out_path, period_formatted, period)
        print("✓ File uploaded to Noloco documents table successfully")
    except Exception as e:
        print(f"\n⚠️  Warning: Failed to upload to Noloco documents: {str(e)}")
        print("  The export file was saved successfully, but the document upload failed.")
    
    print("Done.")


if __name__ == "__main__":
    try:
        run_export()
    except KeyboardInterrupt:
        print("\nInterrupted.")
        exit(130)
    except Exception as e:
        print(f"\nError: {e}")
        import traceback

        traceback.print_exc()
        exit(1)
